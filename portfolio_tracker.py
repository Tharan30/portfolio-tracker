import yfinance as yf
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

def get_price(ticker):
    data = yf.download(ticker, period="1d", interval="1d", progress=False, auto_adjust=True)['Close']
    if not data.empty:
        return float(data.iloc[0].item())
    return None

def append_to_excel(tickers, filename):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Fetch prices
    prices = {ticker: get_price(ticker) for ticker in tickers}
    row = [now] + list(prices.values())

    try:
        # Try to load existing workbook
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        # Write as new row without deleting old data
        df_new = pd.DataFrame([row], columns=["Date"] + tickers)
        df_new.to_excel(writer, sheet_name="Sheet1", index=False, header=False, startrow=book.active.max_row)
        writer.close()

    except FileNotFoundError:
        # If file doesn't exist, create with header
        df_new = pd.DataFrame([row], columns=["Date"] + tickers)
        df_new.to_excel(filename, index=False)

    print(f"Row appended to {filename}")

if __name__ == "__main__":
    tickers = ['SUZLON.NS', 'ETERNAL.NS', 'TATAMOTORS.NS']
    filename = "portfolio.xlsx"
    append_to_excel(tickers, filename)
