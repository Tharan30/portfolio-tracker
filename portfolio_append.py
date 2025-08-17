import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime

def append_to_excel(tickers, filename):
    today = datetime.now().strftime("%Y-%m-%d")  # only date part
    prices = {}

    for ticker in tickers:
        stock = yf.Ticker(ticker)
        info = stock.history(period="1d")
        if not info.empty:
            last_price = info["Close"].iloc[-1]
            prices[ticker] = last_price

    # New row with today's data
    df = pd.DataFrame([{ "Date": today, **prices }])

    if os.path.exists(filename):
        book = load_workbook(filename)
        sheet = book.active

        # Find if today's date already exists in column A
        dates = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)]
        if today in dates:
            row_index = dates.index(today) + 2  # offset for header
            # Update prices in that row
            for col_idx, ticker in enumerate(prices.keys(), start=2):
                sheet.cell(row=row_index, column=col_idx, value=prices[ticker])
        else:
            # Append as new row
            sheet.append(df.iloc[0].tolist())

        book.save(filename)

    else:
        # Create new workbook
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False)

if __name__ == "__main__":
    tickers = ["SUZLON.NS", "TATAMOTORS.NS", "ETERNAL.NS"]
    filename = "portfolio-updates.xlsx"
    append_to_excel(tickers, filename)
